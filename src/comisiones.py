#!/usr/bin/env python3
"""
Calculadora de Comisiones de Ventas — Buho Media  v5.0 (Streamlit Cloud + Drive)
================================================================================
Módulo importable (sin dependencia de terminal). La lógica de negocio v4.7 se
conserva intacta; esta versión añade:

  · Capa de datos Google Drive: descarga los 4 Excel a memoria vía Service Account.
  · Capa de validación estricta: columnas obligatorias por fuente (DataValidationError).
  · Selección explícita de columnas por nombre (sin `iloc`).
  · Salidas en `io.BytesIO` (sin escribir a disco) para `st.download_button`.
  · Entrada única `run_pipeline(fecha_inicio, fecha_fin, sources) -> PipelineResult`.

Fuentes de datos (carpeta compartida de Drive):
  · Margen_Tarifario (.xlsx)
      - Hoja 'Margen_Tarifario': catálogo WMS/Fulfillment (primario)
      - Hoja 'Wing':             catálogo Wing (transición)
  · Catalogo BWMS (.xlsx)        → fallback WMS si no está en Margen_Tarifario
  · Concentrado_Facturas (.xlsx) → facturas emitidas (montos, IVA)
  · Asiento contable (.xlsx)     → validación de pago (Estado + Fecha)

Políticas:
  · Sales Lifecycle: Transition=1%, Hunter=por plan, Farmer=1%,
    Executive Sponsor/Whale=VIP split 50/50
  · Custom/sin plan → tasa Plan D
  · Cobranza: 0-30d=100%, 31-60d=70%, 61-90d=30%, >90d=0%
  · Pago adelantado (días negativos) = factor 1.0
  · Comisiones sobre Total facturado / 1.16 (sin IVA)
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime

import pandas as pd
import streamlit as st
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ╔════════════════════════════════════════════════════════════════════════════╗
# ║              CONFIGURACIÓN DE FUENTES  (ÚNICO PUNTO DE CAMBIO)               ║
# ╚════════════════════════════════════════════════════════════════════════════╝

FACTOR_IVA = 1.16

# Tamaño máximo soportado por archivo (cinturón de seguridad para la subida).
MAX_FILE_MB = 25
MAX_FILE_BYTES = MAX_FILE_MB * 1024 * 1024

# Cada fuente lógica → prefijo de nombre esperado en la carpeta de Drive.
# Cambiar la convención de nombres = editar SOLO este diccionario.
FILE_PATTERNS: dict[str, str] = {
    "catalogo":    "Margen_Tarifario",
    "bwms":        "Catalogo BWMS",
    "concentrado": "Concentrado_Facturas",
    "pagos":       "Asiento contable",
}

# Nombre canónico (fijo) con el que vive cada insumo en Drive. El uploader normaliza a
# estos nombres → cada mes se ACTUALIZA el mismo archivo (posible en Drive personal).
CANONICAL_NAMES: dict[str, str] = {
    "catalogo":    "Margen_Tarifario.xlsx",
    "bwms":        "Catalogo BWMS.xlsx",
    "concentrado": "Concentrado_Facturas.xlsx",
    "pagos":       "Asiento contable.xlsx",
}

# Lectura + escritura: la app descarga las fuentes y también permite subirlas a Drive.
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive"]

XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ════════════════════════════════════════════════════════════════════════════════
#  POLÍTICA DE COMISIONES
# ════════════════════════════════════════════════════════════════════════════════

TRANSITION_RATE       = 0.01
HIGH_VOLUME_THRESHOLD = 150_000
HUNTER_MAX_MONTHS     = 12
STANDARD_PLANS        = {'A', 'B', 'C', 'D'}
VALID_LIFECYCLES      = {'Hunter', 'Farmer', 'Whale', 'Executive Sponsor', 'Transition'}

WING_RATES = {
    'hunter_inbound_base': {'A': 0.0150, 'B': 0.0135, 'C': 0.0115, 'D': 0.0100},
    'hunter_inbound_high': {'A': 0.0250, 'B': 0.0250, 'C': 0.0250, 'D': 0.0250},
    'hunter_outbound':     {'A': 0.0250, 'B': 0.0235, 'C': 0.0215, 'D': 0.0200},
    'farmer':               0.0100,
}
WMS_RATES = {
    'hunter_inbound_base': {'A': 0.0250, 'B': 0.0235, 'C': 0.0215, 'D': 0.0200},
    'hunter_inbound_high': {'A': 0.0250, 'B': 0.0250, 'C': 0.0250, 'D': 0.0250},
    'hunter_outbound':     {'A': 0.0350, 'B': 0.0335, 'C': 0.0315, 'D': 0.0300},
    'farmer':               0.0100,
}
SERVICE_RATES = {'WING': WING_RATES, 'WMS': WMS_RATES}

COBRANZA_BANDAS = [
    (0,  30,  1.00, '0-30 días  (puntual)'),
    (31, 60,  0.70, '31-60 días (70%)'),
    (61, 90,  0.30, '61-90 días (30%)'),
    (91, None, 0.00, '+90 días   (sin comisión)'),
]

# ════════════════════════════════════════════════════════════════════════════════
#  CAPA DE VALIDACIÓN
# ════════════════════════════════════════════════════════════════════════════════


class DataValidationError(Exception):
    """Error de negocio con mensaje amigable, pensado para mostrar en la UI."""


# Columnas obligatorias por fuente (nombres físicos verificados).
REQUIRED_COLUMNS: dict[str, list[str]] = {
    "catalogo":    ['Cliente', 'KAM', 'Plan Comisiones', 'Mes Inicio', 'Año Inicio'],
    "bwms":        ['Cliente', 'KAM', 'Plan Comisiones', 'Mes Inicio', 'Año Inicio'],
    "concentrado": ['Número', 'Estado', 'Estado del CFDI', 'Estado en el SAT', 'Total firmado'],
    "pagos":       ['Número', 'Estado', 'Fecha', 'Facturas'],
}

# Mapeo explícito "campo interno → nombre de columna físico" (reemplaza los iloc).
CONCENTRADO_COLS: dict[str, str] = {
    "invoice_number":    "Número",
    "partner":           "Nombre del partner para mostrar en la factura.",
    "fecha_factura":     "Fecha de factura",
    "fecha_vencimiento": "Fecha de vencimiento",
    "estado":            "Estado",
    "estado_cfdi":       "Estado del CFDI",
    "estado_sat":        "Estado en el SAT",
    "total_facturado":   "Total firmado",
}
ASIENTO_COLS: dict[str, str] = {
    "invoice_number": "Facturas",
    "estado_pago":    "Estado",
    "fecha_pago":     "Fecha",
}


def validate_columns(df: pd.DataFrame, clave: str, nombre_amigable: str) -> None:
    """Valida que `df` no esté vacío y tenga las columnas de REQUIRED_COLUMNS[clave]."""
    if df.empty:
        raise DataValidationError(f"El archivo '{nombre_amigable}' está vacío.")
    faltantes = [c for c in REQUIRED_COLUMNS[clave] if c not in df.columns]
    if faltantes:
        raise DataValidationError(
            f"A '{nombre_amigable}' le faltan columnas obligatorias: {faltantes}.")


def _select_columns(df: pd.DataFrame, cols_map: dict[str, str],
                    nombre_amigable: str) -> pd.DataFrame:
    """Selecciona por nombre y renombra a los campos internos; error amigable si falta una."""
    faltantes = [col for col in cols_map.values() if col not in df.columns]
    if faltantes:
        raise DataValidationError(
            f"A '{nombre_amigable}' le faltan columnas obligatorias: {faltantes}.")
    inverso = {fisico: interno for interno, fisico in cols_map.items()}
    return df[list(cols_map.values())].rename(columns=inverso)


# ════════════════════════════════════════════════════════════════════════════════
#  CAPA DE DATOS — GOOGLE DRIVE
# ════════════════════════════════════════════════════════════════════════════════


@st.cache_resource(show_spinner=False)
def get_drive_service():
    """Cliente de Google Drive autenticado con la Service Account de st.secrets.

    Cacheado como recurso (una sola inicialización por sesión de servidor).
    """
    creds = Credentials.from_service_account_info(
        dict(st.secrets["gcp_service_account"]), scopes=DRIVE_SCOPES)
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _is_valid_excel(name: str) -> bool:
    """Descarta temporales de Excel ('~$…') y exige extensión .xlsx real."""
    return name.endswith(".xlsx") and not name.startswith("~$")


def _list_folder(folder_id: str) -> list[dict]:
    """Lista los archivos (no en papelera) de la carpeta, del más reciente al más antiguo.

    Los parámetros `*AllDrives` hacen que funcione tanto con carpetas de "Mi unidad"
    como de Unidades compartidas (Shared Drives) sin configuración extra.
    """
    service = get_drive_service()
    resp = service.files().list(
        q=f"'{folder_id}' in parents and trashed = false",
        fields="files(id, name, modifiedTime)",
        orderBy="modifiedTime desc",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    return resp.get("files", [])


def _resolve_file_ids(folder_id: str) -> dict[str, dict]:
    """Empareja cada FILE_PATTERN con el .xlsx más reciente y válido de la carpeta."""
    archivos = [f for f in _list_folder(folder_id) if _is_valid_excel(f["name"])]
    resueltos: dict[str, dict] = {}
    for clave, prefijo in FILE_PATTERNS.items():
        match = next((f for f in archivos if f["name"].startswith(prefijo)), None)
        if match is None:
            raise DataValidationError(
                f"No se encontró en Drive un archivo .xlsx que empiece con '{prefijo}'.")
        resueltos[clave] = match
    return resueltos


@st.cache_data(show_spinner=False)
def _download_bytes(file_id: str, modified: str) -> bytes:
    """Descarga un archivo de Drive a bytes.

    Cache-key = (file_id, modified). Si el archivo cambia en Drive su `modifiedTime`
    cambia y la caché se invalida sola, evitando re-descargas innecesarias por rerun.
    """
    service = get_drive_service()
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(
        buffer, service.files().get_media(fileId=file_id, supportsAllDrives=True))
    done = False
    while not done:
        # num_retries: reintentos con backoff exponencial ante cortes de red
        # transitorios (p. ej. Errno 10053 en Windows por firewall/antivirus).
        _, done = downloader.next_chunk(num_retries=5)
    return buffer.getvalue()


@dataclass
class DataSources:
    """Contenido crudo (bytes) de los 4 Excel. Los loaders envuelven cada campo en
    un `io.BytesIO` fresco por lectura (evita problemas de puntero al leer 2 hojas)."""
    catalogo: bytes
    bwms: bytes
    concentrado: bytes
    pagos: bytes


@st.cache_data(show_spinner="Descargando archivos desde Google Drive…")
def load_data_sources_from_drive(folder_id: str) -> DataSources:
    """Resuelve los 4 archivos de la carpeta y los descarga a memoria."""
    resueltos = _resolve_file_ids(folder_id)
    return DataSources(**{
        clave: _download_bytes(meta["id"], meta["modifiedTime"])
        for clave, meta in resueltos.items()
    })


def clear_drive_cache() -> None:
    """Invalida las cachés de descarga tras subir/actualizar archivos en Drive."""
    load_data_sources_from_drive.clear()
    _download_bytes.clear()


# ── Subida de fuentes a Drive (administración) ─────────────────────────────────

def match_prefix(filename: str) -> str | None:
    """Devuelve la clave de FILE_PATTERNS cuyo prefijo coincide con `filename`, o None."""
    if not _is_valid_excel(filename):
        return None
    for clave, prefijo in FILE_PATTERNS.items():
        if filename.startswith(prefijo):
            return clave
    return None


def _find_file_id_by_name(folder_id: str, filename: str, service) -> str | None:
    """Busca un archivo por nombre exacto dentro de la carpeta (para hacer upsert)."""
    safe_name = filename.replace("\\", "\\\\").replace("'", "\\'")
    resp = service.files().list(
        q=f"'{folder_id}' in parents and name = '{safe_name}' and trashed = false",
        fields="files(id)", supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    files = resp.get("files", [])
    return files[0]["id"] if files else None


def upload_file_to_drive(folder_id: str, filename: str, content: bytes) -> str:
    """Sube un .xlsx a la carpeta (upsert por nombre). Devuelve 'creado' o 'actualizado'.

    Nota: en Drive personal ("Mi unidad") la cuenta de servicio NO puede *crear*
    archivos nuevos (no tiene cuota). Sí puede *actualizar* uno existente que sea tuyo.
    Si el `create` falla por cuota, se lanza un DataValidationError accionable.
    """
    service = get_drive_service()
    media = MediaIoBaseUpload(io.BytesIO(content), mimetype=XLSX_MIMETYPE, resumable=False)
    existing_id = _find_file_id_by_name(folder_id, filename, service)
    if existing_id:
        service.files().update(
            fileId=existing_id, media_body=media, supportsAllDrives=True).execute()
        return "actualizado"
    try:
        service.files().create(
            body={"name": filename, "parents": [folder_id]},
            media_body=media, fields="id", supportsAllDrives=True).execute()
        return "creado"
    except HttpError as exc:
        if "storageQuotaExceeded" in str(exc):
            raise DataValidationError(
                f"No se puede crear '{filename}' con la cuenta de servicio en un Drive "
                "personal (las SA no tienen almacenamiento propio). Sube este archivo "
                "por primera vez de forma manual a la carpeta; después la app podrá "
                "actualizarlo automáticamente. Para creación 100% desde la app, usa una "
                "Unidad compartida (Shared Drive)."
            ) from exc
        raise


# ════════════════════════════════════════════════════════════════════════════════
#  UTILIDADES
# ════════════════════════════════════════════════════════════════════════════════

def parse_date(text: str) -> date:
    text = text.strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d'):
        try: return datetime.strptime(text, fmt).date()
        except ValueError: pass
    raise ValueError(f"Fecha '{text}' no reconocida. Use DD/MM/AAAA.")

def cobranza_factor(days_late: int) -> float:
    if days_late <= 0: return 1.0
    for _, hi, f, _ in COBRANZA_BANDAS:
        if hi is None or days_late <= hi: return f
    return 0.0

def cobranza_tramo(days_late: int) -> str:
    if days_late <= 0: return 'Pago adelantado'
    for _, hi, _, label in COBRANZA_BANDAS:
        if hi is None or days_late <= hi: return label
    return '+90 días (sin comisión)'

def account_age_months(mes: int, año: int, ref: date) -> int:
    return max(1, (ref.year - int(año)) * 12 + (ref.month - int(mes)) + 1)

def _normalize_plan(val: str) -> str:
    """Normaliza plan: 'Plan A' → 'A', 'custom' → 'D', NaN → 'D'."""
    v = str(val).strip().upper()
    if v in STANDARD_PLANS: return v
    m = re.search(r'PLAN\s+([A-D])', v)
    if m: return m.group(1)
    return 'D'  # Custom / desconocido → Plan D

def _normalize_tipo_venta(val: str) -> str:
    """Normaliza: 'Outbound' → 'Outbound Sale', 'Inbound' → 'Inbound Sale'."""
    v = str(val).strip()
    if v.upper().startswith('OUTBOUND'): return 'Outbound Sale'
    if v.upper().startswith('INBOUND'):  return 'Inbound Sale'
    return v

def _is_valid_vendedor(v: str) -> bool:
    """Filtra nombres que no son vendedores (nombres de empresa en columna KAM)."""
    v = str(v).strip()
    if v.upper() in ('NAN', 'NONE', '', 'NAT'): return False
    if len(v.split()) > 4: return False  # Nombres de empresa suelen ser largos
    return True

# ════════════════════════════════════════════════════════════════════════════════
#  CARGA DE CATÁLOGOS
# ════════════════════════════════════════════════════════════════════════════════

def _normalize_lifecycle(val: str) -> str:
    v = str(val).strip()
    if v in VALID_LIFECYCLES: return v
    return 'Transition'

def load_catalog_main(source: bytes) -> pd.DataFrame:
    """Carga hoja 'Margen_Tarifario' (catálogo WMS primario) desde bytes."""
    raw = pd.read_excel(io.BytesIO(source), sheet_name='Margen_Tarifario')
    raw.columns = raw.columns.astype(str).str.strip()
    validate_columns(raw, 'catalogo', 'Margen Tarifario')

    df = pd.DataFrame({
        'Name':            raw.get('Cliente', pd.Series(dtype=str)).astype(str).str.strip(),
        'Vendedor':        raw.get('KAM', pd.Series(dtype=str)).astype(str).str.strip(),
        'Plan':            raw.get('Plan Comisiones', 'D').astype(str).apply(_normalize_plan),
        'Tipo de Venta':   raw.get('Tipo de Venta', 'Outbound Sale').astype(str).apply(_normalize_tipo_venta),
        'Mes Inicio':      pd.to_numeric(raw.get('Mes Inicio', 1), errors='coerce').fillna(1).astype(int),
        'Año Inicio':      pd.to_numeric(raw.get('Año Inicio', 2024), errors='coerce').fillna(2024).astype(int),
        'Nombre Fiscal':   raw.get('Nombre Fiscal', pd.Series(dtype=str)).astype(str).str.strip(),
        'Nombre2':         raw.get('Partner/Nombre', pd.Series(dtype=str)).astype(str).str.strip(),
        'Margen':          pd.to_numeric(raw.get('Margen', 0.30), errors='coerce').fillna(0.30),
        'Canal':           'WMS',
        'Sales Lifecycle': raw.get('Sales Lifecycle', 'Transition').astype(str).apply(_normalize_lifecycle),
        'Farmer':          raw.get('Farmer', pd.Series(dtype=str)).astype(str).str.strip(),
        'Handoff':         raw.get('Handoff', pd.Series(dtype=str)).astype(str).str.strip(),
        '_source':         'MT1',
    })
    df = df[df['Vendedor'].apply(_is_valid_vendedor)].copy()
    return df.reset_index(drop=True)


def load_catalog_wing(source: bytes) -> pd.DataFrame:
    """Carga hoja 'Wing' de Margen_Tarifario (esquema propio de transición)."""
    df = pd.read_excel(io.BytesIO(source), sheet_name='Wing')
    df.columns = df.columns.astype(str).str.strip()
    for c in ['Name', 'Vendedor', 'Canal']:
        if c in df.columns: df[c] = df[c].astype(str).str.strip()
    df['Tipo de Venta']   = df.get('Tipo de Venta', 'Outbound Sale').astype(str).apply(_normalize_tipo_venta)
    df['Plan']            = df.get('Plan', 'D').astype(str).apply(_normalize_plan)
    df['Mes Inicio']      = pd.to_numeric(df.get('Mes Inicio', 1), errors='coerce').fillna(1).astype(int)
    df['Año Inicio']      = pd.to_numeric(df.get('Año Inicio', 2024), errors='coerce').fillna(2024).astype(int)
    df['Nombre Fiscal']   = df.get('Nombre Fiscal', pd.Series(dtype=str)).astype(str).str.strip()
    df['Nombre2']         = df.get('Nombre2', pd.Series(dtype=str)).astype(str).str.strip()
    df['Margen']          = pd.to_numeric(df.get('Margen', 0.30), errors='coerce').fillna(0.30)
    df['Canal']           = df.get('Canal', 'WING').astype(str).str.strip().str.upper()
    df['Canal']           = df['Canal'].where(df['Canal'].isin(['WING', 'WMS']), 'WING')
    df['Sales Lifecycle'] = df.get('Sales Lifecycle', 'Transition').astype(str).apply(_normalize_lifecycle)
    df['Farmer']          = df.get('Farmer', pd.Series(dtype=str)).astype(str).str.strip()
    df['Handoff']         = df.get('Handoff', pd.Series(dtype=str)).astype(str).str.strip()
    df['_source']         = 'Wing'
    df['Vendedor']        = df['Vendedor'].str.title()
    df = df[~df['Vendedor'].isin(['nan', 'None', '', 'NaN', 'Nan'])].copy()
    return df.reset_index(drop=True)


def load_catalog_bwms(source: bytes) -> pd.DataFrame:
    """Carga hoja 'clientes' de Catalogo BWMS (fallback) desde bytes."""
    raw = pd.read_excel(io.BytesIO(source), sheet_name='clientes')
    raw.columns = raw.columns.astype(str).str.strip()
    validate_columns(raw, 'bwms', 'Catálogo BWMS')
    df = pd.DataFrame({
        'Name':            raw['Cliente'].astype(str).str.strip(),
        'Vendedor':        raw['KAM'].astype(str).str.strip(),
        'Plan':            raw.get('Plan Comisiones', 'D').astype(str).apply(_normalize_plan),
        'Tipo de Venta':   raw.get('Tipo de Venta', 'Outbound Sale').astype(str).apply(_normalize_tipo_venta),
        'Mes Inicio':      pd.to_numeric(raw.get('Mes Inicio', 1), errors='coerce').fillna(1).astype(int),
        'Año Inicio':      pd.to_numeric(raw.get('Año Inicio', 2024), errors='coerce').fillna(2024).astype(int),
        'Nombre Fiscal':   raw.get('Nombre Fiscal', pd.Series(dtype=str)).astype(str).str.strip(),
        'Nombre2':         raw.get('Partner/Nombre', pd.Series(dtype=str)).astype(str).str.strip(),
        'Margen':          pd.to_numeric(raw.get('Margen', 0.30), errors='coerce').fillna(0.30),
        'Canal':           'WMS',
        'Sales Lifecycle': raw.get('Sales Lifecycle', 'Transition').astype(str).apply(_normalize_lifecycle),
        'Farmer':          raw.get('Farmer', pd.Series(dtype=str)).astype(str).str.strip(),
        '_source':         'BWMS',
    })
    df = df[~df['Vendedor'].isin(['nan', 'None', '', 'NaN'])].copy()
    return df.reset_index(drop=True)


def merge_all_catalogs(mt1: pd.DataFrame, bwms: pd.DataFrame, wing: pd.DataFrame) -> pd.DataFrame:
    """Merge con prioridad: MT1 > BWMS > Wing. Sin duplicados por nombre."""
    combined = mt1.copy()
    seen = set()
    for col in ('Name', 'Nombre Fiscal', 'Nombre2'):
        if col in combined.columns:
            seen |= set(combined[col].dropna().astype(str).str.strip().str.upper()) - {'NAN','NONE',''}

    for source_df in [bwms, wing]:
        keep = []
        for _, row in source_df.iterrows():
            keys = {str(row.get(c,'')).strip().upper()
                    for c in ('Name','Nombre Fiscal','Nombre2')} - {'NAN','NONE',''}
            if not keys & seen:
                keep.append(row)
                seen |= keys
        if keep:
            combined = pd.concat([combined, pd.DataFrame(keep)], ignore_index=True)

    return combined.reset_index(drop=True)


# ════════════════════════════════════════════════════════════════════════════════
#  CARGA DE FACTURAS
# ════════════════════════════════════════════════════════════════════════════════

def load_invoices(concentrado: bytes, pagos: bytes,
                  start: date, end: date) -> pd.DataFrame:
    """
    Concentrado = facturas con montos. Asiento = validación de pago + fecha.
    Base Comisionable = Total facturado / 1.16.
    Días de atraso = fecha_pago − fecha_factura (negativo = adelantado).
    Selección de columnas por nombre (CONCENTRADO_COLS / ASIENTO_COLS), sin iloc.
    """
    conc_raw = pd.read_excel(io.BytesIO(concentrado), sheet_name=0)
    conc_raw.columns = conc_raw.columns.astype(str).str.strip()
    validate_columns(conc_raw, 'concentrado', 'Concentrado de Facturas')
    conc = _select_columns(conc_raw, CONCENTRADO_COLS, 'Concentrado de Facturas')

    conc_df = pd.DataFrame({
        'invoice_number':    conc['invoice_number'].astype(str).str.strip(),
        'partner':           conc['partner'].astype(str).str.strip().str.upper(),
        'fecha_factura':     pd.to_datetime(conc['fecha_factura'], errors='coerce'),
        'fecha_vencimiento': pd.to_datetime(conc['fecha_vencimiento'], errors='coerce'),
        'estado':            conc['estado'].astype(str).str.strip(),
        'estado_cfdi':       conc['estado_cfdi'].astype(str).str.strip(),
        'estado_sat':        conc['estado_sat'].astype(str).str.strip(),
        'total_facturado':   pd.to_numeric(conc['total_facturado'], errors='coerce').fillna(0),
    })
    # Solo facturas fiscalmente vigentes (registradas, CFDI firmado y validado ante SAT).
    conc_df = conc_df[
        (conc_df['estado'] == 'Registrado') &
        (conc_df['estado_cfdi'] == 'Firmado') &
        (conc_df['estado_sat'] == 'Validado')
    ].copy()

    asiento_raw = pd.read_excel(io.BytesIO(pagos), sheet_name='Sheet1')
    asiento_raw.columns = asiento_raw.columns.astype(str).str.strip()
    validate_columns(asiento_raw, 'pagos', 'Asiento Contable')
    asiento = _select_columns(asiento_raw, ASIENTO_COLS, 'Asiento Contable')

    pago = pd.DataFrame({
        'invoice_number': asiento['invoice_number'].astype(str).str.strip(),
        'estado_pago':    asiento['estado_pago'].astype(str).str.strip(),
        'fecha_pago':     pd.to_datetime(asiento['fecha_pago'], errors='coerce'),
    })
    pago = pago[pago['estado_pago'] == 'Pagado'].copy()
    # Una factura puede tener varios asientos; nos quedamos con la última fecha de pago.
    pago = pago.groupby('invoice_number', as_index=False).agg(fecha_pago=('fecha_pago', 'max'))

    merged = conc_df.merge(pago, on='invoice_number', how='inner')
    # Corte del período por FECHA DE PAGO (no de factura).
    merged = merged[
        (merged['fecha_pago'] >= pd.Timestamp(start)) &
        (merged['fecha_pago'] <= pd.Timestamp(end).replace(hour=23, minute=59, second=59))
    ].copy()

    merged['base_comisionable'] = (merged['total_facturado'] / FACTOR_IVA).round(2)
    merged['days_late'] = merged.apply(
        lambda r: (r['fecha_pago'] - r['fecha_factura']).days
        if pd.notna(r['fecha_pago']) and pd.notna(r['fecha_factura']) else 0, axis=1)

    return merged[['partner','invoice_number','fecha_factura','fecha_pago',
                    'total_facturado','base_comisionable','days_late']].copy()


def get_payment_date_range(sources: DataSources) -> tuple[date, date] | None:
    """Rango (min, max) de las fechas de pago 'Pagado' del Asiento contable.

    Sirve para que la UI muestre el período disponible y acote el selector de fechas.
    Devuelve None si no hay fechas de pago válidas.
    """
    asiento_raw = pd.read_excel(io.BytesIO(sources.pagos), sheet_name='Sheet1')
    asiento_raw.columns = asiento_raw.columns.astype(str).str.strip()
    validate_columns(asiento_raw, 'pagos', 'Asiento Contable')
    asiento = _select_columns(asiento_raw, ASIENTO_COLS, 'Asiento Contable')

    pagados = asiento[asiento['estado_pago'].astype(str).str.strip() == 'Pagado']
    fechas = pd.to_datetime(pagados['fecha_pago'], errors='coerce').dropna()
    if fechas.empty:
        return None
    return fechas.min().date(), fechas.max().date()


# ════════════════════════════════════════════════════════════════════════════════
#  ÍNDICE + BÚSQUEDA
# ════════════════════════════════════════════════════════════════════════════════

def build_index(catalog: pd.DataFrame) -> dict:
    idx = {}
    for _, row in catalog.sort_values(['Año Inicio', 'Mes Inicio']).iterrows():
        for col in ('Nombre2', 'Nombre Fiscal', 'Name'):
            val = str(row.get(col, '')).strip().upper()
            if val and val not in ('NAN', 'NONE', ''):
                idx[val] = row
    return idx

def find_client(name: str, idx: dict):
    return idx.get(name.strip().upper())


# ════════════════════════════════════════════════════════════════════════════════
#  COMISIONES
# ════════════════════════════════════════════════════════════════════════════════

def commission_rate(plan: str, tipo_venta: str, canal: str,
                    lifecycle: str, total_amount: float = 0, age: int = 1,
                    is_handoff: bool = False):
    rates = SERVICE_RATES.get(canal, WING_RATES)

    if lifecycle in ('Whale', 'Executive Sponsor'):
        return ('Executive Sponsor', rates['farmer'], 'Executive Sponsor')

    if lifecycle == 'Farmer':
        return ('Farmer', rates['farmer'], 'Farmer')

    # Transition = no clasificado explícitamente → determinar por antigüedad
    # Si ≤12 meses → tratar como Hunter (usar tasas por plan/tipo)
    # Si >12 meses → tratar como Farmer (1%)
    if lifecycle == 'Transition':
        if age > HUNTER_MAX_MONTHS:
            return ('Farmer', rates['farmer'], f'Farmer (mes {age})')
        # else: continuar con lógica Hunter abajo

    # Hunter (explícito o Transition con ≤12 meses)
    if tipo_venta == 'Outbound Sale':
        r = rates['hunter_outbound'][plan]
        return ('Hunter', r, f'Hunter Outbound Plan {plan}')

    if total_amount > HIGH_VOLUME_THRESHOLD:
        r = rates['hunter_inbound_high'][plan]
        return ('Hunter', r, f'Hunter Inbound >{HIGH_VOLUME_THRESHOLD:,.0f} MXN')
    if age <= 6:
        r = rates['hunter_inbound_base'][plan]
        return ('Hunter', r, f'Hunter Inbound base mes {age}')
    if is_handoff:
        return ('Hunter', 0.0, f'Hunter Inbound Handoff mes {age} (0%)')
    r = rates['hunter_inbound_base'][plan]
    return ('Hunter', r, f'Hunter Inbound mes {age}')


def _make_row(vendedor, period_label, client, base_comisionable, monto_cobrado,
              tasa, rol, note, days_late, factor, tramo, canal) -> dict:
    base_comm = round(base_comisionable * tasa, 2)
    adj_comm  = round(base_comm * factor, 2)
    return {
        'Vendedor':              vendedor,
        'Período':               period_label,
        'Cliente':               client['Name'],
        'Nombre Fiscal':         str(client.get('Nombre Fiscal', '')),
        'Plan':                  str(client.get('Plan', 'D')).upper(),
        'Tipo de Venta':         str(client.get('Tipo de Venta', '')),
        'Canal':                 canal,
        'Sales Lifecycle':       str(client.get('Sales Lifecycle', 'Transition')),
        'Mes de Cuenta':         account_age_months(client['Mes Inicio'], client['Año Inicio'], date.today()),
        'Monto Cobrado MXN':     monto_cobrado,
        'Base Comisionable MXN': round(base_comisionable, 2),
        'Tasa':                  tasa,
        'Comisión Base MXN':     base_comm,
        'Días de Atraso':        days_late,
        'Tramo Cobranza':        tramo,
        'Factor Cobranza':       factor,
        'Comisión Ajustada MXN': adj_comm,
        'Notas':                 note,
    }


def build_invoice_row(client, base_comisionable: float, total_facturado: float,
                      total_base_period: float, period_label: str, ref: date,
                      days_late_val: int) -> list:
    plan      = str(client.get('Plan', 'D')).upper()
    tipo      = str(client.get('Tipo de Venta', 'Outbound Sale'))
    canal     = str(client.get('Canal', 'WMS')).upper()
    lifecycle = str(client.get('Sales Lifecycle', 'Transition'))
    vendedor  = str(client.get('Vendedor', '—'))
    handoff   = str(client.get('Handoff', '')).strip().upper() not in ('', 'NAN', 'NONE')
    age       = account_age_months(client['Mes Inicio'], client['Año Inicio'], ref)

    rol, tasa, note = commission_rate(plan, tipo, canal, lifecycle, total_base_period, age, handoff)

    factor = cobranza_factor(days_late_val)
    tramo  = cobranza_tramo(days_late_val)

    row = _make_row(vendedor, period_label, client,
                    base_comisionable, total_facturado,
                    tasa, rol, note, days_late_val, factor, tramo, canal)
    row['Mes de Cuenta'] = age
    return [row]


# ════════════════════════════════════════════════════════════════════════════════
#  EXCEL
# ════════════════════════════════════════════════════════════════════════════════

HEADER_BG = '1F4E79'; ALT_BG = 'D9E1F2'; TOTAL_BG = 'FFF2CC'

def _header_style(ws):
    hf = PatternFill('solid', fgColor=HEADER_BG)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = hf
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

def _apply_rows(ws, df):
    alt = PatternFill('solid', fgColor=ALT_BG)
    cols = list(df.columns)
    money = {i+1 for i,c in enumerate(cols) if any(k in str(c) for k in ('MXN','Comisión','Impacto','Comisionable','Cobrado'))}
    pct   = {i+1 for i,c in enumerate(cols) if any(k in str(c) for k in ('Tasa','Efectiva','Factor'))}
    for ri, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        is_total = isinstance(row[0].value, str) and '⬛' in row[0].value
        for cell in row:
            if is_total:
                cell.fill = PatternFill('solid', fgColor=HEADER_BG)
                cell.font = Font(bold=True, color='FFFFFF', size=9)
            elif ri % 2 == 0: cell.fill = alt
            if cell.column in money:
                cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal='right')
            elif cell.column in pct:
                cell.number_format = '0.00%'; cell.alignment = Alignment(horizontal='right')

def _auto_width(ws, df):
    for i, col in enumerate(df.columns, 1):
        w = max(len(str(col)), *(len(str(v)) for v in df[col].dropna().astype(str)))
        ws.column_dimensions[get_column_letter(i)].width = min(w + 3, 55)

def write_sheet(writer, df, sheet_name):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    _header_style(ws); _apply_rows(ws, df); _auto_width(ws, df)


# ════════════════════════════════════════════════════════════════════════════════
#  REPORTES POR VENDEDOR / GLOBAL  (en memoria → io.BytesIO)
# ════════════════════════════════════════════════════════════════════════════════

def _build_por_cliente(detail_df, start, end):
    if detail_df.empty: return pd.DataFrame()
    grp = ['Cliente','Nombre Fiscal','Plan','Tipo de Venta','Canal','Sales Lifecycle','Mes de Cuenta']
    agg = (detail_df.groupby(grp, as_index=False)
           .agg(Facturas=('Base Comisionable MXN','count'),
                Monto_Cobrado=('Monto Cobrado MXN','sum'),
                Base_Com=('Base Comisionable MXN','sum'),
                Com_Base=('Comisión Base MXN','sum'),
                Com_Ajust=('Comisión Ajustada MXN','sum'))
           .rename(columns={'Monto_Cobrado':'Monto Cobrado MXN','Base_Com':'Base Comisionable MXN',
                            'Com_Base':'Comisión Base MXN','Com_Ajust':'Comisión Ajustada MXN'}))
    agg['Tasa'] = agg.apply(lambda r: r['Comisión Base MXN']/r['Base Comisionable MXN'] if r['Base Comisionable MXN'] else 0, axis=1)

    tramo_m, factor_m = {}, {}
    for cli, g in detail_df.groupby('Cliente'):
        bs, aj = g['Comisión Base MXN'].sum(), g['Comisión Ajustada MXN'].sum()
        factor_m[cli] = aj/bs if bs else 1.0
        tm = g.groupby('Tramo Cobranza')['Base Comisionable MXN'].sum()
        dom = tm.idxmax() if not tm.empty else '—'
        if (tm > 0).sum() > 1:
            punt = tm.get('Pago adelantado', 0) + tm.get('0-30 días  (puntual)', 0)
            tramo_m[cli] = f'Mixto ({punt/tm.sum():.0%} puntual)'
        else: tramo_m[cli] = dom

    agg['Tramo Cobranza'] = agg['Cliente'].map(tramo_m).fillna('—')
    agg['Factor Cobranza'] = agg['Cliente'].map(factor_m).fillna(1.0)
    agg['Impacto Cobranza MXN'] = agg['Comisión Base MXN'] - agg['Comisión Ajustada MXN']
    agg = agg.sort_values('Comisión Ajustada MXN', ascending=False).reset_index(drop=True)
    agg.insert(0, 'Fecha Inicio', f"{start:%d/%m/%Y}")
    agg.insert(1, 'Fecha Fin', f"{end:%d/%m/%Y}")
    cols = ['Fecha Inicio','Fecha Fin','Cliente','Nombre Fiscal','Plan','Tipo de Venta','Canal',
            'Sales Lifecycle','Mes de Cuenta','Facturas','Monto Cobrado MXN',
            'Base Comisionable MXN','Tasa','Comisión Base MXN','Tramo Cobranza',
            'Factor Cobranza','Comisión Ajustada MXN','Impacto Cobranza MXN']
    return agg[[c for c in cols if c in agg.columns]]


def write_vendedor_excel(vendedor: str, df: pd.DataFrame,
                         start: date, end: date) -> tuple[io.BytesIO, str, float, float]:
    """Genera el reporte individual de un vendedor en memoria.

    Devuelve (buffer, nombre_archivo, base_total, comisión_total).
    """
    safe = re.sub(r'[^\w\s-]', '', vendedor).strip().replace(' ', '_')
    filename = f"Comisiones_{safe}_{start:%d-%m-%Y}_{end:%d-%m-%Y}.xlsx"
    pc = _build_por_cliente(df, start, end)
    base_t = df['Base Comisionable MXN'].sum()
    comm_b = df['Comisión Base MXN'].sum()
    comm_a = df['Comisión Ajustada MXN'].sum()
    impacto = comm_b - comm_a
    lc = df.groupby('Sales Lifecycle')['Cliente'].nunique()
    rows = [
        ('Vendedor', vendedor), ('Período', f"{start:%d/%m/%Y} — {end:%d/%m/%Y}"), ('',''),
        ('Clientes activos', df['Cliente'].nunique()),
        ('  — Hunter', lc.get('Hunter',0)), ('  — Farmer', lc.get('Farmer',0)),
        ('  — Executive Sponsor', lc.get('Executive Sponsor',0)),
        ('  — Transition', lc.get('Transition',0)), ('',''),
        (f'Base Comisionable MXN (÷{FACTOR_IVA})', base_t),
        ('Comisión base MXN', comm_b), ('Impacto cobranza MXN', impacto),
        ('TOTAL COMISIÓN MXN', comm_a),
    ]
    rdf = pd.DataFrame(rows, columns=['Concepto','Valor'])

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as w:
        write_sheet(w, rdf, 'Resumen'); write_sheet(w, pc, 'Por Cliente')
        ws = w.sheets['Resumen']
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
            rn, vc = r[0].row, r[1]
            if rn in {10,11,12,13} and isinstance(vc.value, (int,float)):
                vc.number_format = '#,##0.00'
                if rn == 13:
                    vc.font = Font(bold=True, size=11); vc.fill = PatternFill('solid', fgColor=TOTAL_BG)
                    r[0].font = Font(bold=True, size=11); r[0].fill = PatternFill('solid', fgColor=TOTAL_BG)
                elif rn == 12: vc.font = Font(color='CC0000', size=9)
    buffer.seek(0)
    return buffer, filename, base_t, comm_a


def write_global_summary(summary_rows: list, unmatched: set,
                         start: date, end: date) -> tuple[io.BytesIO, str]:
    """Genera el Resumen Global en memoria. Devuelve (buffer, nombre_archivo)."""
    filename = f"Resumen_Global_{start:%d-%m-%Y}_{end:%d-%m-%Y}.xlsx"
    sdf = pd.DataFrame(summary_rows).sort_values('Comisión MXN', ascending=False)
    tc, tk = sdf['Base Comisionable MXN'].sum(), sdf['Comisión MXN'].sum()
    total = {'Vendedor':'⬛ TOTAL','Clientes':sdf['Clientes'].sum(),
             'Base Comisionable MXN':tc,'Comisión MXN':tk,
             'Tasa Efectiva':tk/tc if tc else 0,'Archivo':''}
    sdf = pd.concat([sdf, pd.DataFrame([total])], ignore_index=True)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as w:
        write_sheet(w, sdf, 'Resumen por Vendedor')
        ws = w.sheets['Resumen por Vendedor']
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor=HEADER_BG)
        if unmatched:
            write_sheet(w, pd.DataFrame({'Sin match en Concentrado': sorted(unmatched)}), 'Sin Match')
    buffer.seek(0)
    return buffer, filename


# ════════════════════════════════════════════════════════════════════════════════
#  PIPELINE  (entrada única para la UI)
# ════════════════════════════════════════════════════════════════════════════════

@dataclass
class PipelineResult:
    """Resultado listo para consumir por la UI (métricas + buffers de descarga)."""
    total_facturas: int
    base_comisionable_global: float
    total_comisiones: float
    periodo: str
    vendedores: list[str]
    resumen_global: tuple[io.BytesIO, str]
    reportes_vendedor: dict[str, tuple[io.BytesIO, str]]
    sin_match: set[str]


def run_pipeline(fecha_inicio: date, fecha_fin: date,
                 sources: DataSources) -> PipelineResult:
    """Orquesta catálogos → facturas → comisiones → reportes en memoria.

    No imprime ni escribe a disco: todo se devuelve en un PipelineResult.
    Lanza DataValidationError si una fuente no cumple el esquema esperado.
    """
    if fecha_fin < fecha_inicio:
        raise DataValidationError("La fecha fin no puede ser anterior a la fecha inicio.")

    period = f"{fecha_inicio:%d/%m/%Y} – {fecha_fin:%d/%m/%Y}"

    # ── Catálogos ─────────────────────────────────────────────────────────────
    cat_mt1  = load_catalog_main(sources.catalogo)
    cat_wing = load_catalog_wing(sources.catalogo)
    cat_bwms = load_catalog_bwms(sources.bwms)
    catalog  = merge_all_catalogs(cat_mt1, cat_bwms, cat_wing)
    cat_idx  = build_index(catalog)

    # ── Facturas con pago validado ────────────────────────────────────────────
    invoices = load_invoices(sources.concentrado, sources.pagos, fecha_inicio, fecha_fin)
    ptotals = (invoices.groupby('partner')['base_comisionable'].sum().to_dict()
               if not invoices.empty else {})

    # ── Comisiones (una fila por factura) ─────────────────────────────────────
    all_rows: list[dict] = []
    unmatched: set[str] = set()
    for _, inv in invoices.iterrows():
        client = find_client(inv['partner'], cat_idx)
        if client is None:
            unmatched.add(inv['partner'])
            continue
        tbp = ptotals.get(inv['partner'], inv['base_comisionable'])
        all_rows.extend(build_invoice_row(
            client, inv['base_comisionable'], inv['total_facturado'],
            tbp, period, fecha_fin, inv['days_late']))

    if not all_rows:
        raise DataValidationError(
            "No se generaron comisiones para el período seleccionado "
            "(sin facturas pagadas con match en catálogo).")

    detail = pd.DataFrame(all_rows)

    # ── Reportes por vendedor + resumen ───────────────────────────────────────
    vendedores = sorted(v for v in detail['Vendedor'].unique()
                        if v not in ('nan', 'None', '—', ''))
    reportes_vendedor: dict[str, tuple[io.BytesIO, str]] = {}
    summary: list[dict] = []
    for v in vendedores:
        vdf = detail[detail['Vendedor'] == v].copy()
        buffer, filename, base, com = write_vendedor_excel(v, vdf, fecha_inicio, fecha_fin)
        reportes_vendedor[v] = (buffer, filename)
        summary.append({
            'Vendedor': v, 'Clientes': vdf['Cliente'].nunique(),
            'Base Comisionable MXN': round(base, 2), 'Comisión MXN': round(com, 2),
            'Tasa Efectiva': com/base if base else 0, 'Archivo': filename,
        })

    resumen_global = write_global_summary(summary, unmatched, fecha_inicio, fecha_fin)

    return PipelineResult(
        total_facturas=len(invoices),
        base_comisionable_global=sum(r['Base Comisionable MXN'] for r in summary),
        total_comisiones=sum(r['Comisión MXN'] for r in summary),
        periodo=period,
        vendedores=vendedores,
        resumen_global=resumen_global,
        reportes_vendedor=reportes_vendedor,
        sin_match=unmatched,
    )
